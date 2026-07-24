#!/usr/bin/env python3
"""Lector de archivos para el sistema de presupuestos.

Lee un archivo (.xlsx / .csv / .pdf), detecta una tabla de items y la
devuelve como JSON por stdout. Si no puede mapear columnas claras, devuelve
el texto crudo para que lo interprete la IA del backend.

Salida (una línea JSON):
  {"source":"table","items":[{name,quantity,unit,unit_price,detail}, ...]}
  {"source":"text","text":"..."}
  {"error":"..."}

Uso:  python extract.py <archivo> <ext>
"""
import sys
import json
import csv
import re
import unicodedata

# Forzar stdout a UTF-8 (en Windows por defecto usa cp1252 y rompe las tildes)
try:
    sys.stdout.reconfigure(encoding='utf-8')
except Exception:
    pass


def norm(s):
    """Normaliza un encabezado: minúsculas, sin tildes, solo letras/números."""
    s = str(s or '').strip().lower()
    s = unicodedata.normalize('NFKD', s)
    s = ''.join(c for c in s if not unicodedata.combining(c))  # saca tildes (ó->o, ñ->n)
    return re.sub(r'[^a-z0-9]', '', s)


def to_num(v):
    """Convierte texto de número (formato es-AR o en-US) a float."""
    if v is None:
        return 0.0
    s = str(v).strip()
    if not s:
        return 0.0
    s = re.sub(r'[^\d,.\-]', '', s)
    if not s or s in ('-', '.', ','):
        return 0.0
    if ',' in s and '.' in s:
        # 1.234.567,89  ->  1234567.89
        s = s.replace('.', '').replace(',', '.')
    elif ',' in s:
        # 1234,89  ->  1234.89  (la coma es decimal en AR)
        s = s.replace(',', '.')
    try:
        return float(s)
    except ValueError:
        return 0.0


def _first(hn, exact=(), contains=()):
    """Devuelve el índice de la 1ra columna que matchea (exacto primero)."""
    for i, h in enumerate(hn):
        if h in exact:
            return i
    for i, h in enumerate(hn):
        if any(k in h for k in contains):
            return i
    return None


def find_desc(hn):
    # Columnas fuertes de descripción primero (evita pegar con "Item" o "Precio Items")
    idx = _first(hn,
                 exact=('descripcion', 'detalle', 'concepto', 'rubro', 'tarea',
                        'trabajo', 'producto', 'articulo', 'material', 'nombre', 'concepto'),
                 contains=('descripc', 'detalle', 'concepto', 'trabajo', 'producto', 'articulo'))
    if idx is not None:
        return idx
    # Último recurso: una columna llamada exactamente "item"
    for i, h in enumerate(hn):
        if h == 'item':
            return i
    return None


def find_qty(hn):
    return _first(hn, exact=('cantidad', 'cant', 'qty', 'cantidades'), contains=('cantidad',))


def find_unit(hn):
    idx = _first(hn, exact=('unidad', 'unid', 'umedida', 'medida', 'um'),
                 contains=('unidad', 'medida'))
    if idx is not None:
        return idx
    for i, h in enumerate(hn):
        if h == 'un':
            return i
    return None


PRICE_UNIT_KEYS = ('preciounit', 'preciounitario', 'preciouni', 'punit',
                   'valorunitario', 'valorunit', 'unitario', 'preciodeunidad')
PRICE_TOTAL_KEYS = ('importe', 'total', 'subtotal', 'montoitem', 'preciototal',
                    'precioitem', 'precioitems', 'importetotal')

# Unidad pegada al número en la propia celda de cantidad (ej "0,18 m3", "42 ml").
# Nota: se busca sobre el texto ya normalizado (norm_keep_spaces), donde "m³"/"m²"
# ya quedaron como "m3"/"m2" (el NFKD decompone los superíndices a dígitos normales).
UNIT_TOKEN_RE = re.compile(
    r'(?:^|\s)(m3|m2|ml|mts?|kg|lt|hs|un|gl|dias?)(?=\s|$)',
    re.IGNORECASE,
)
UNIT_TOKEN_MAP = {
    'm3': 'm³', 'm2': 'm²',
    'ml': 'm', 'mt': 'm', 'mts': 'm', 'kg': 'kg', 'lt': 'lt',
    'hs': 'hs', 'un': 'un.', 'gl': 'gl', 'dia': 'día', 'dias': 'días',
}


def find_price(hn):
    """Devuelve (índice, es_total). "es_total" marca columnas tipo Importe/Total
    (precio ya multiplicado por la cantidad) para no confundirlas con el precio unitario."""
    idx = _first(hn, exact=PRICE_UNIT_KEYS)
    if idx is not None:
        return idx, False
    idx = _first(hn, contains=('preciounit', 'unitario'))
    if idx is not None:
        return idx, False
    idx = _first(hn, exact=('precio', 'valor', 'costo'))
    if idx is not None:
        return idx, False
    idx = _first(hn, exact=PRICE_TOTAL_KEYS)
    if idx is not None:
        return idx, True
    idx = _first(hn, contains=('importe', 'total', 'subtotal'))
    if idx is not None:
        return idx, True
    idx = _first(hn, contains=('precio', 'costo', 'valor'))
    if idx is not None:
        return idx, False
    return None, False


def split_qty_unit(raw):
    """Si la celda de cantidad trae la unidad pegada ("0,18 m3"), separa ambas.
    Ojo: hay que sacar el token de la unidad ANTES de convertir a número, porque
    unidades como "m3"/"m2" tienen un dígito propio que si no se contaminaría
    con la cantidad (ej "0,18 m3" -> "0,183" si no se lo saca antes)."""
    s = str(raw or '').strip()
    m = UNIT_TOKEN_RE.search(norm_keep_spaces(s))
    if not m:
        return to_num(s), None
    unit = UNIT_TOKEN_MAP.get(m.group(1).lower())
    number_part = (s[:m.start()] + s[m.end():]).strip()
    return to_num(number_part), unit


def norm_keep_spaces(s):
    s = unicodedata.normalize('NFKD', s.lower())
    return ''.join(c for c in s if not unicodedata.combining(c))


def rows_to_items(rows):
    """Convierte filas (lista de listas; la 1ra con encabezados) en items."""
    rows = [r for r in rows if r and any(str(c or '').strip() for c in r)]
    if not rows:
        return None

    # Buscamos la fila de encabezados: la primera con descripción + (precio o cantidad).
    # Ventana amplia (30 filas) porque muchas planillas reales tienen membrete,
    # título y renglones en blanco antes de la tabla en sí.
    header_idx = None
    for i, r in enumerate(rows[:30]):
        hn = [norm(c) for c in r]
        pi_probe, _ = find_price(hn)
        if find_desc(hn) is not None and (pi_probe is not None or find_qty(hn) is not None):
            header_idx = i
            break
    if header_idx is None:
        return None

    hn = [norm(c) for c in rows[header_idx]]
    di = find_desc(hn)
    qi = find_qty(hn)
    ui = find_unit(hn)
    pi, pi_is_total = find_price(hn)
    if di is None:
        return None

    items = []
    for r in rows[header_idx + 1:]:
        def cell(idx):
            if idx is None or idx >= len(r) or r[idx] is None:
                return ''
            return str(r[idx]).strip()

        name = cell(di)
        # saltar vacíos y filas que repiten el encabezado
        if not name or norm(name) in ('descripcion', 'detalle', 'item', 'concepto', 'rubro'):
            continue

        qty_raw = cell(qi)
        quantity, unit_from_qty = split_qty_unit(qty_raw) if qty_raw else (0, None)
        quantity = quantity or 1

        unit_cell = cell(ui)
        # Si no hay columna de unidad separada (o vino vacía), usamos la que se
        # coló pegada al número de cantidad (ej "0,18 m3"); si no hay ninguna, "un.".
        unit = unit_cell[:20] if unit_cell else (unit_from_qty or 'un.')

        raw_price = to_num(cell(pi))
        # Si la columna de precio detectada es un TOTAL (Importe/Total, no "unitario"),
        # el precio por unidad real es total ÷ cantidad — evita cargar el total como
        # si fuera el precio de 1 sola unidad.
        unit_price = (raw_price / quantity) if (pi_is_total and quantity) else raw_price

        item = {
            'name': name[:200],
            'quantity': quantity,
            'unit': unit,
            'unit_price': round(unit_price, 2),
            'detail': '',
        }
        items.append(item)
    return items or None


def read_xlsx_sheets(path):
    """Devuelve las filas de CADA hoja del archivo (no solo la activa) — algunas
    planillas reales tienen la tabla de items en una hoja que no es la primera."""
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    sheets = []
    for ws in wb.worksheets:
        rows = [['' if c is None else c for c in row] for row in ws.iter_rows(values_only=True)]
        if rows:
            sheets.append(rows)
    return sheets


def read_csv(path):
    with open(path, 'r', encoding='utf-8-sig', errors='replace', newline='') as f:
        sample = f.read(4096)
        f.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=',;\t|')
        except Exception:
            dialect = csv.excel
        return [list(r) for r in csv.reader(f, dialect)]


def read_pdf(path):
    import pdfplumber
    table_rows = []
    text_parts = []
    with pdfplumber.open(path) as pdf:
        for page in pdf.pages:
            for t in (page.extract_tables() or []):
                table_rows.extend(t)
            txt = page.extract_text() or ''
            if txt:
                text_parts.append(txt)
    return table_rows, '\n'.join(text_parts)


def rows_to_text(rows):
    return '\n'.join('\t'.join('' if c is None else str(c) for c in r) for r in rows)


def looks_like_descriptions(items):
    """True si los nombres parecen descripciones reales (no números/códigos cortos).
    Atrapa el caso de PDFs donde la columna detectada termina siendo el Nº de fila."""
    if not items:
        return False
    good = sum(1 for it in items
               if len(it['name'].strip()) > 3 and not re.fullmatch(r'[\d.,$%\s]+', it['name'].strip()))
    return good >= max(1, len(items) // 2)


def best_result(rows):
    """Mapea a items si la tabla es clara; si no, devuelve el texto para la IA."""
    items = rows_to_items(rows)
    if items and looks_like_descriptions(items):
        return {'source': 'table', 'items': items[:300]}
    return {'source': 'text', 'text': rows_to_text(rows)[:6000]}


def best_result_multi(sheets):
    """Prueba cada hoja del Excel y usa la primera que arme una tabla clara.
    Si ninguna lo logra, junta el texto de todas para que lo interprete la IA."""
    for rows in sheets:
        res = best_result(rows)
        if res['source'] == 'table':
            return res
    combined = '\n\n'.join(rows_to_text(rows) for rows in sheets)
    return {'source': 'text', 'text': combined[:6000]}


def main():
    if len(sys.argv) < 3:
        print(json.dumps({'error': 'uso: extract.py <archivo> <ext>'}))
        return
    path, ext = sys.argv[1], sys.argv[2].lower()
    try:
        if ext == 'xlsx':
            print(json.dumps(best_result_multi(read_xlsx_sheets(path)), ensure_ascii=False))
            return
        if ext == 'csv':
            print(json.dumps(best_result(read_csv(path)), ensure_ascii=False))
            return
        if ext == 'pdf':
            table_rows, text = read_pdf(path)
            res = best_result(table_rows) if table_rows else None
            if res and res['source'] == 'table':
                print(json.dumps(res, ensure_ascii=False))
            else:
                # texto del PDF (más confiable que la tabla cruda) para que lo arme la IA
                print(json.dumps({'source': 'text', 'text': text[:6000]}, ensure_ascii=False))
            return
        print(json.dumps({'error': f'Formato no soportado: {ext}'}))
    except Exception as e:
        print(json.dumps({'error': f'No se pudo leer el archivo: {e}'}, ensure_ascii=False))


if __name__ == '__main__':
    main()
